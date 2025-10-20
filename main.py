import os
import sys

import numpy as np
import pandas as pd
import xgboost as xgb
#from imblearn.over_sampling import SMOTE
from sklearn import svm, neighbors
from sklearn.calibration import CalibratedClassifierCV
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score, f1_score, confusion_matrix
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder, StandardScaler
from anonymize import l_diversity
from explainability import shap_explainability
from explainability import lime_explainability
from explainability import lime_explainability_change
from tabulate import tabulate
from sklearn.model_selection import cross_validate, StratifiedKFold
import openpyxl
from openpyxl.styles import Font
import random
from openpyxl.utils import get_column_letter

sys.path.insert(0, '/Users/lucacena/Desktop/Tesi/Prove_codice/Prova_anonypy/Anonypy_Second_trial/anonypy')

results_df = pd.DataFrame({
    "Dataset": pd.Series(dtype='str'),
    "Modello": pd.Series(dtype='str'),
    "Anonimizzato": pd.Series(dtype='str'),
    "F1 Score": pd.Series(dtype='float'),
    "Accuracy": pd.Series(dtype='float'),
    "k": pd.Series(dtype='int'),
    "l": pd.Series(dtype='int')
})


#Funzione che permette di concatenare e aggiungere risultati ottenuti
def log_result(df, dataset, modello, anonimizzato, f1, accuracy, k, l):
    new_row = {
        "Dataset": dataset,
        "Modello": modello,
        "Anonimizzato": anonimizzato,
        "F1 Score": round(f1, 3),
        "Accuracy": round(accuracy, 3),
        "k": k,
        "l": l
    }
    df.loc[len(df)] = new_row
    return df


def preprocess_student(filename):
    df = pd.read_csv(filename)
    keep_cols = [
        "Age", "Course", "Gender", "CGPA", "Stress_Level", "Depression_Score", "Anxiety_Score",
        "Sleep_Quality", "Physical_Activity", "Diet_Quality", "Relationship_Status",
        "Counseling_Service_Use", "Extracurricular_Involvement", "Semester_Credit_Load", "Residence_Type"
    ]
    df = df[keep_cols]

    def max_mental_issue(row):
        values = {
            1: row["Stress_Level"],
            2: row["Depression_Score"],
            3: row["Anxiety_Score"]
        }
        if all(v <= 2 for v in values.values()):
            return 0  # none
        max_value = max(values.values())
        max_keys = [k for k, v in values.items() if v == max_value]
        return random.choice(max_keys)

    # Creazione della nuova colonna
    df["mental_illness"] = df.apply(max_mental_issue, axis=1)
    df['Mental_Severity_Score'] = df['Stress_Level'] + df['Depression_Score'] + df['Anxiety_Score']
    # Gestione valori mancanti in CGPA
    df["CGPA"] = df["CGPA"].fillna(df["CGPA"].median())
    # df['has_mental_illness'] = (df['mental_illness'] > 0).astype(int)
    # df["stress_depression_diff"] = (df["Stress_Level"] - df["Depression_Score"]).abs()
    df["stress_depression_diff"] = df["Stress_Level"] - df["Depression_Score"]
    df["stress_anxiety_diff"] = (df["Stress_Level"] - df["Anxiety_Score"]).abs()
    df.drop(columns=["Stress_Level", "Depression_Score", "Anxiety_Score"], inplace=True)
    # Salva il nuovo dataset
    df.to_csv("dataset/students_mental_health_processed.csv", index=False)

    return df


def load_file(data_name):
    if data_name == 'stroke':
        filename = 'dataset/healthcare-dataset-stroke-data.csv'
        target = "stroke"
        categorical_columns = ['gender', 'ever_married', 'work_type', 'Residence_type', 'smoking_status']
    elif data_name == "adult":
        filename = 'dataset/adult.csv'
        categorical_columns = ['gender', 'race', 'marital-status', 'education', 'native-country', 'relationship',
                               'workclass', 'occupation', 'income']
        target = 'income'
    elif data_name == 'diabetes':
        filename = 'dataset/diabetes.csv'
        categorical_columns = ['gender', 'smoking_history']
        target = 'diabetes'
    elif data_name == 'mental_health':
        filename = 'dataset/students_mental_health.csv'
        df = preprocess_student(filename)
        categorical_columns = ['Course', 'Gender', 'Sleep_Quality', 'Physical_Activity', 'Diet_Quality',
                               'Relationship_Status', 'Counseling_Service_Use',
                               'Extracurricular_Involvement', 'Residence_Type']
        target = 'mental_illness'
    else:
        filename = 'dataset/NHANES_age_prediction.csv'
        target = "age_group"
        categorical_columns = ['age_group']
    if data_name not in ['mental_health', 'depression', 'thyroid', 'health']:
        df = pd.read_csv(filename)
    #print(df.dtypes)
    #print(df.head())

    if data_name == 'stroke':
        df = df.drop('bmi', axis=1)
    elif data_name == "NHANES":
        df = df.drop('SEQN', axis=1)
        print("Lunghezza:", len(df))
    return df, target, categorical_columns


def label_encode_columns(data, to_encode):
    le = LabelEncoder()
    for col in to_encode:
        data[col] = data[col].astype('category')
        data[col] = le.fit_transform(data[col])


def split_data(dataset, target):
    X = dataset.drop(target, axis=1)
    Y = dataset[target]
    return X, Y


def fit_model(model_name, X, Y):
    train_features, test_features, train_labels, test_labels = train_test_split(X, Y, test_size=0.20,
                                                                                random_state=42)
    test_features = pd.DataFrame(test_features, columns=X.columns)
    if model_name == 'RF':
        model = RandomForestClassifier(random_state=42)
    elif model_name == 'SVM':
        svm_mod = svm.LinearSVC(dual=False)
        model = CalibratedClassifierCV(svm_mod, ensemble=False)
    elif model_name == 'KNN':
        model = neighbors.KNeighborsClassifier(5, weights='distance')
    elif model_name == 'XGB':
        if dataset_name == 'mental_health':
            model = xgb.XGBClassifier(objective="multi:softmax", num_class=4, random_state=42,
                                      enable_categorical=True)
        #elif dataset_name in ['thyroid', 'health']:
        #elif dataset_name in ['health', 'NHANES']:
        #    model = xgb.XGBClassifier(objective="multi:softmax", num_class=3, random_state=42,
        #                              enable_categorical=True)
        else:
            model = xgb.XGBClassifier(objective="binary:hinge", random_state=42, enable_categorical=True)
    else:
        raise ValueError('Invalid model type')
    model.fit(train_features, train_labels)
    predictions = model.predict(test_features)

    return model, predictions, test_features


def fit_model_lime(model_name, X, Y):
    # Salvo i row_id (se presenti)
    if "_row_id" in X.columns:
        row_ids = X["_row_id"].copy()
        X = X.drop(columns=["_row_id"])
    else:
        row_ids = pd.Series(X.index, name="_row_id")
    # Split mantenendo anche row_id
    X_train, X_test, y_train, y_test, rid_train, rid_test = train_test_split(
        X, Y, row_ids, test_size=0.20, random_state=42
    )
    # Ricostruisco i DataFrame con row_id incluso → utile per LIME
    train_features = X_train.copy()
    train_features["_row_id"] = rid_train.values
    test_features = X_test.copy()
    test_features["_row_id"] = rid_test.values
    feature_cols = X_train.columns.tolist()
    # Ora il modello vede solo le feature (senza _row_id)
    X_train_model = X_train
    X_test_model = X_test
    #print(f"[fit_model_lime] Training set shape (senza _row_id): {X_train_model.shape}")
    #print(f"[fit_model_lime] Test set shape (senza _row_id): {X_test_model.shape}")
    #print(f"[fit_model_lime] Numero di feature usate per il modello: {X_train_model.shape[1]}")
    #print(f"Test features: {test_features.shape[1]}")
    if model_name == 'RF':
        model = RandomForestClassifier(random_state=42)
    elif model_name == 'SVM':
        svm_mod = svm.LinearSVC(dual=False)
        model = CalibratedClassifierCV(svm_mod, ensemble=False)
    elif model_name == 'KNN':
        model = neighbors.KNeighborsClassifier(5, weights='distance')
    elif model_name == 'XGB':
        if dataset_name == 'mental_health':
            model = xgb.XGBClassifier(objective="multi:softmax", num_class=4, random_state=42,
                                      enable_categorical=True)
        else:
            model = xgb.XGBClassifier(objective="binary:hinge", random_state=42, enable_categorical=True)
    else:
        raise ValueError('Invalid model type')

    model.fit(X_train_model, y_train)
    predictions = model.predict(X_test_model)

    # Restituisco anche i test_features con _row_id incluso
    return model, predictions, test_features, feature_cols, train_features


def predictions(dataset_name, X, Y, model_type, anon, k, l):
    global results_df
    if anon:
        anonimizzato = "anonimizzato"
    else:
        anonimizzato = "non anonimizzato"
    skf = StratifiedKFold(n_splits=10, shuffle=True)
    if model_type == 'XGB':
        for col in X.columns:
            if X[col].dtype == 'object':
                X[col] = X[col].astype('category')
    # Modello base
    if model_type == 'RF':
        base_model = RandomForestClassifier(random_state=42)
    elif model_type == 'SVM':
        svm_mod = svm.LinearSVC(dual=False)
        base_model = CalibratedClassifierCV(svm_mod, ensemble=False)
    elif model_type == 'KNN':
        base_model = neighbors.KNeighborsClassifier(5, weights='distance')
    elif model_type == 'XGB':
        if dataset_name in ['mental_health', 'depression']:
            base_model = xgb.XGBClassifier(objective="multi:softmax", num_class=4, random_state=42,
                                           enable_categorical=True)
        #elif dataset_name in ['thyroid', 'health', 'NHANES']:
        #    base_model = xgb.XGBClassifier(objective="multi:softmax", num_class=3, random_state=42,
        #                                   enable_categorical=True)
        else:
            base_model = xgb.XGBClassifier(objective="binary:hinge", random_state=42, enable_categorical=True)
    else:
        raise ValueError('Invalid model type')
    # Cross-validation
    cv_results = cross_validate(base_model, X, Y, cv=skf,
                                scoring=['accuracy', 'f1_weighted'],
                                return_train_score=False)
    acc_mean = cv_results['test_accuracy'].mean()
    #f1_mean = cv_results['test_f1_macro'].mean()
    #f1_mean = cv_results['test_f1_micro'].mean()
    f1_mean = cv_results['test_f1_weighted'].mean()
    results_df = log_result(results_df, dataset_name, model_type,
                            anonimizzato, f1_mean, acc_mean, k, l)


if __name__ == '__main__':
    model_name = ['RF', 'SVM', 'KNN', 'XGB']
    data_set = ['mental_health']
    #data_set = ['mental_health', 'stroke', 'NHANES', 'diabetes', 'adult']
    #data_set = ['NHANES', 'adult', 'diabetes']
    for dataset_name in data_set:
        dataset, target, categorical_columns = load_file(dataset_name)
        for col in categorical_columns:
            dataset[col] = dataset[col].astype("category")
        if 'id' in dataset.columns:
            dataset = dataset.drop('id', axis=1)
        dataset_lime = dataset.copy()
        dataset_lime["_row_id"] = dataset_lime.index.astype(str)
        #print(f"Colonne orignali: {list(dataset.columns)}")
        k_values = [2, 5, 10, 15, 20, 50, 70, 100]
        if dataset_name == 'mental_health':
            l_values = range(2, 5)
        elif dataset_name == 'NHANES':
            l_values = range(2, 4)
        else:
            l_values = range(2, 3)
        k_max = max(k_values)
        l_max = max(l_values)
        # Dataset originale → encoding prima di passarlo ai modelli
        dataset_encoded = dataset.copy()
        dataset_lime_encoded = dataset_lime.copy()
        if categorical_columns:
            label_encode_columns(dataset_encoded, categorical_columns)
        if categorical_columns:
            label_encode_columns(dataset_lime_encoded, categorical_columns)
        os.makedirs("dataset_orig_encod", exist_ok=True)
        dataset_encoded.to_csv(os.path.join("dataset_orig_encod", f"{dataset_name}_encoded.csv"), index=False)
        for model in model_name:
            print(f"Dataset: {dataset_name}, Modello: {model}, Anonimizzato: no")
            #X, Y = split_data(dataset_encoded, target)
            X, Y = split_data(dataset_lime_encoded, target)
            #print(f"Dimesnione X: {list(X.columns)}")
            #predictions(dataset_name, X, Y, model, 0, 0, 0)
            #model_fitted, pred, test_features = fit_model(model, X, Y)
            model_fitted, pred, test_features, feature_cols, train_features_orig = fit_model_lime(model, X, Y)
            #problematic = analyze_model_predictions(model_fitted, test_features, target, dataset_encoded)
            #shap_explainability(model_fitted, test_features, 0, dataset_name, model, 0, 0, make_beeswarm=True)
            #lime_explainability(dataset_encoded, dataset_name, test_features, model_fitted, 0, model, 0, 0, target)
            for k in k_values:
                for l in l_values:
                    if k < l:
                        continue
                    print(f"\n--- {dataset_name}: k={k}, l={l} ---")
                    # Anonimizzazione (ancora con categorie testuali)
                    dataset_anonymized = l_diversity(dataset_name, dataset.copy(), k, l)
                    dataset_anonymized_lime = dataset_anonymized.copy()
                    #print(f"Colonne anon: {list(dataset_anonymized_lime.columns)}")
                    dataset_anonymized = dataset_anonymized.drop(columns=["_row_id"], errors="ignore")
                    dataset_anonymized_encoded = dataset_anonymized.copy()
                    dataset_anonymized_lime_encoded = dataset_anonymized_lime.copy()
                    if categorical_columns:
                        label_encode_columns(dataset_anonymized_encoded, categorical_columns)
                    if categorical_columns:
                        label_encode_columns(dataset_anonymized_lime_encoded, categorical_columns)
                    #print(f"Colonne anon dopo encoded: {list(dataset_anonymized_lime_encoded.columns)}")
                    # Salvo i dataset anonimizzati numerici
                    dataset_anonymized_encoded.to_csv(
                        f"dataset_anon/{dataset_name}_k{k}_l{l}.csv", sep='\t', index=False
                    )
                    # Modelli
                    #print(f"Dataset: {dataset_name}, Modello: {model}, Anonimizzato: sì")
                    #X_anon, Y_anon = split_data(dataset_anonymized_encoded, target)
                    X_anon, Y_anon = split_data(dataset_anonymized_lime_encoded, target)
                    #print(f"Dimesnione X_anon: {list(X_anon.columns)}")
                    #predictions(dataset_name, X_anon, Y_anon, model, 1, k, l)
                    #model_fitted_anon, pred_anon, test_features_anon = fit_model(model, X_anon, Y_anon)
                    #print("Lunghezza feat_cols prima:", len(feature_cols))
                    model_fitted_anon, pred_anon, test_features_anon, feature_cols, train_features_anon = fit_model_lime(model, X_anon, Y_anon)
                    #print("Lunghezza feat_cols dopo :", len(feature_cols))
                    """if k == k_max and l == l_max:
                        shap_explainability(model_fitted_anon, test_features_anon, k, dataset_name, model, 1, l,
                                            make_beeswarm=True)
                    else:
                        shap_explainability(model_fitted_anon, test_features_anon, k, dataset_name, model, 1, l,
                                            make_beeswarm=False)"""
                    #lime_explainability(dataset_anonymized_encoded, dataset_name, test_features_anon, model_fitted_anon,1, model, k, l, target)
                    lime_explainability_change(train_features_orig, dataset_name, test_features, test_features_anon, model_fitted, model_fitted_anon, pred, pred_anon, k, l, train_features_anon, model, feature_cols)
                    #lime_explainability_change(dataset_anonymized_encoded, dataset_name, test_features,test_features_anon, model_fitted, model_fitted_anon, model, k, l, target)
                    #validation_results = lime_explainability_2(dataset_anonymized_encoded, dataset_name, test_features_anon, model_fitted_anon, 1, model, k, l, target)

    results_df.to_csv("risultati_classificatori.csv", index=False)
    # Salvataggio unico Excel formattato
    excel_path = "risultati_classificatori_formattato.xlsx"
    results_df.to_excel(excel_path, index=False)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(excel_path)
    print(f"Tabella Excel formattata salvata come: {excel_path}")

    data3 = {
        "ID": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12],
        "Età": [29, 53, 30, 42, 45, 44, 51, 52, 28, 31, 89, 94],
        "ZIPCODE": [47677, 47602, 47678, 47905, 47909, 47906, 47605, 47673, 46607, 47607, 58892, 56738],
        "CAP": [10001, 10001, 10001, 10002, 10002, 10002, 10003, 10003, 10003, 10008, 10009, 10009],
        "Sesso": ["M", "F", "M", "F", "F", "M", "M", "F", "F", "F", "M", "M"],
        "Malattia": ["Diabete", "Diabete", "Ipertensione", "Diabete", "Cancro", "Diabete", "Ipertensione", "Asma",
                     "Asma",
                     "Cancro", "Prostata", "Prostata"]
    }

    categorical3 = {"Sesso", "Malattia"}
    df3 = pd.DataFrame(data=data3)

    for name in categorical3:
        df3[name] = df3[name].astype("category")

    #label_encode_columns(df3, categorical3)

    # df2["Malattia"] = df2["Malattia"] + "-" + df2["ID"].astype(str)
    print(df3)
    df3 = l_diversity("prova", df3, 4, 3)
    print(df3)
    #label_encode_columns(df3, categorical3)
    #print(df3)
